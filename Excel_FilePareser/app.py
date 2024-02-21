from flask import Flask ,render_template ,  request , jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import  Marshmallow
from openpyxl import load_workbook


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///data.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app) 
ma = Marshmallow(app)

class File_Data(db.Model):
    id = db.Column(db.Integer(),primary_key=True)
    name = db.Column(db.String(100),nullable=False)
    age = db.Column(db.String(100),nullable=False)
    country = db.Column(db.String(100),nullable=False)

    def __init__(self,name,age,country):
        self.name = name
        self.age = age
        self.country = country


with app.app_context():
    db.create_all()

class File_schema(ma.Schema):
    class Meta:
        fields = ['id','name','age','country']


File_Schema_Single = File_schema()
File_Schema_Multiple = File_schema(many=True)


@app.route('/user' )
def get():
    return render_template('upload.html')

@app.route('/upload', methods=['POST'])
def Data_add():
    if request.method == 'POST':
        data = request.files['file']

        # if not data.filename.endswith(('.csv', '.xls', '.xlsx')):
        #     return "Error: Only CSV, XLS, and XLSX files are allowed to be uploaded."
        
        recieve_data = load_workbook(data)
        loaded_data = recieve_data.active

        for i in loaded_data.iter_rows(min_row=2, values_only=True):
            data = File_Data(name=i[0], age=i[1], country=i[2])
            db.session.add(data)
        db.session.commit()

        return jsonify(message="Data Received")

@app.route('/get/<int:id>', methods=['GET'])
def get_one(id):
    if request.method == 'GET':
        # one_data = File_Data.query.filter_by(id=id).first()
        one_data = File_Data.query.get(id)
        if one_data is not None:
            result = File_Schema_Single.dump(one_data)
            return jsonify({"Data": result})
        else:
            return jsonify({'message': 'ID not found'})



@app.route('/get_all', methods=['GET'])
def get_all():
    if request.method == 'GET':
        all_data = File_Data.query.all()
        result = File_Schema_Multiple.dump(all_data)
        return jsonify({'Data': result})



@app.route('/update/<int:id>', methods=['PUT'])
def update(id):
    if request.method == 'PUT':
        data = File_Data.query.filter_by(id=id).first()
        # data = File_Data.query.get(id)
        if data is None:
            return jsonify({"error": "ID not found"})
        
        data_id = request.json
        required_data = ['name', 'age', 'country']
        for i in required_data:
            if i not in data_id:
                miss_data = [i for i in required_data if i not in data_id]
                return jsonify({"error": f"Missing required fields: {' '.join(miss_data)}"})
      
        data.name = data_id['name']
        data.age = data_id['age']
        data.country = data_id['country']
        db.session.commit()

        updated_data = File_Schema_Single.dump(data)
        return jsonify({"message": "ID updated successfully", "updated_data": updated_data})



@app.route('/delete/<int:id>', methods=['DELETE'])
def delete(id):
    if request.method == 'DELETE':
        data = File_Data.query.get(id)
        # data = File_Data.query.filter_by(id=id).first()
        if data is None:
            return jsonify({"error": "ID not found"})
        db.session.delete(data)
        db.session.commit()
        return jsonify({"message": "ID deleted successfully"})


if __name__ == '__main__':
    app.run(debug=True)
